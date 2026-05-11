import reflex as rx
from typing import TypedDict
import uuid
from datetime import datetime
import io
import openpyxl
from app.shared_store import SharedStore
from app.states.auth_state import AuthState
import logging


def _safe_int(val) -> int:
    """Safely convert Excel cell value to int, returning 0 for errors/non-numeric."""
    if val is None:
        return 0
    try:
        return int(float(val))
    except (ValueError, TypeError):
        return 0


class InventoryState(rx.State):
    active_inventory_tab: str = "dashboard"
    inventory_items: list[dict] = []
    inventory_audits: list[dict] = []
    filtered_items: list[dict] = []
    discrepancy_items: list[dict] = []
    total_items: int = 0
    matched_count: int = 0
    discrepancy_count: int = 0
    accuracy_pct: float = 0.0
    imc_total_rows: int = 0
    imc_unique_refs: int = 0
    latin_total_products: int = 0
    status_chart_data: list[dict] = [
        {"name": "Coincide", "value": 0, "fill": "#22c55e"},
        {"name": "Sobrante", "value": 0, "fill": "#f59e0b"},
        {"name": "Faltante", "value": 0, "fill": "#ef4444"},
    ]
    inventory_search: str = ""
    inventory_filter: str = "all"
    import_message: str = ""
    is_importing: bool = False
    show_discrepancies: bool = False

    def _apply_filters(self):
        q = self.inventory_search.lower().strip()
        f = self.inventory_filter
        result = []
        for item in self.inventory_items:
            if q and q not in str(item.get("product_code", "")).lower():
                continue
            d = item.get("diferencia", 0)
            est = item.get("estado", "")
            if f == "match" and d != 0:
                continue
            if f == "sobrante" and est != "Sobrante":
                continue
            if f == "faltante" and est != "Faltante":
                continue
            result.append(item)
        self.filtered_items = result

    def _compute_stats(self):
        n = len(self.inventory_items)
        self.total_items = n
        m = 0
        sobrante = 0
        faltante = 0
        for item in self.inventory_items:
            d = item.get("diferencia", 0)
            if d == 0:
                m += 1
            elif d > 0:
                sobrante += 1
            else:
                faltante += 1
        self.matched_count = m
        self.discrepancy_count = n - m
        self.accuracy_pct = m / n * 100 if n > 0 else 0.0
        self.status_chart_data = [
            {"name": "Coincide", "value": m, "fill": "#22c55e"},
            {"name": "Sobrante", "value": sobrante, "fill": "#f59e0b"},
            {"name": "Faltante", "value": faltante, "fill": "#ef4444"},
        ]
        self._apply_filters()
        if self.show_discrepancies:
            self.discrepancy_items = [
                i for i in self.inventory_items if i.get("diferencia", 0) != 0
            ]
        else:
            self.discrepancy_items = []

    @rx.event
    def load_inventory_data(self):
        data = SharedStore.load_all_inventory_data()
        self.inventory_items = data.get("inventory_items", [])
        self.inventory_audits = data.get("inventory_audits", [])
        imc_raw = data.get("imc_raw_data", [])
        latin_raw = data.get("latin_raw_data", [])
        self.imc_total_rows = len(imc_raw)
        self.imc_unique_refs = (
            len(set((d.get("referencia", "") for d in imc_raw)))
            if imc_raw
            else 0
        )
        self.latin_total_products = (
            len(set((d.get("product_code", "") for d in latin_raw)))
            if latin_raw
            else 0
        )
        self._compute_stats()

    @rx.event
    def set_inventory_tab(self, tab: str):
        self.active_inventory_tab = tab

    @rx.event
    def set_inventory_search(self, val: str):
        self.inventory_search = val
        self._apply_filters()

    @rx.event
    def set_inventory_filter(self, val: str):
        self.inventory_filter = val
        self._apply_filters()

    @rx.event
    def toggle_discrepancies(self):
        self.show_discrepancies = not self.show_discrepancies
        if self.show_discrepancies:
            self.discrepancy_items = [
                i for i in self.inventory_items if i.get("diferencia", 0) != 0
            ]
        else:
            self.discrepancy_items = []

    @rx.event
    async def handle_inventory_excel_upload(self, files: list[rx.UploadFile]):
        auth = await self.get_state(AuthState)
        if auth.current_role != "admin":
            yield rx.toast("No tienes permisos para importar inventarios.")
            return
        if not files:
            return
        self.is_importing = True
        yield
        try:
            file = files[0]
            upload_data = await file.read()
            upload_dir = rx.get_upload_dir()
            upload_dir.mkdir(parents=True, exist_ok=True)
            original_path = upload_dir / "inventario_original.xlsx"
            with original_path.open("wb") as f:
                f.write(upload_data)
            wb = openpyxl.load_workbook(
                filename=io.BytesIO(upload_data), data_only=True, read_only=True
            )
            if len(wb.sheetnames) < 1:
                self.import_message = "El Excel está vacío."
                wb.close()
                self.is_importing = False
                yield rx.toast(self.import_message)
                return
            imc_sheet = wb.worksheets[0]
            imc_pivot = {}
            imc_count = 0
            imc_raw_list = []
            for row in imc_sheet.iter_rows(min_row=2, values_only=True):
                if row and len(row) > 2 and (row[2] is not None):
                    ref = str(row[2]).strip()
                    if ref:
                        imc_count += 1
                        stock = _safe_int(row[3]) if len(row) > 3 else 0
                        imc_pivot[ref] = imc_pivot.get(ref, 0) + stock
                        imc_raw_list.append(
                            {
                                "row_id": str(uuid.uuid4()),
                                "referencia": ref,
                                "material": str(row[0]).strip()
                                if len(row) > 0 and row[0] is not None
                                else "",
                                "lote": str(row[1]).strip()
                                if len(row) > 1 and row[1] is not None
                                else "",
                                "stock": stock,
                                "ubicacion": str(row[7]).strip()
                                if len(row) > 7 and row[7] is not None
                                else "",
                            }
                        )
            SharedStore.set_imc_raw_data(imc_raw_list)
            self.imc_total_rows = imc_count
            self.imc_unique_refs = len(imc_pivot)
            latin_products = {}
            latin_raw_list = []
            if len(wb.sheetnames) >= 3:
                latin_sheet = wb.worksheets[2]
                for row in latin_sheet.iter_rows(min_row=6, values_only=True):
                    if row and len(row) > 0 and (row[0] is not None):
                        code = str(row[0]).strip()
                        if code and code != "Product Code":
                            stock = _safe_int(row[2]) if len(row) > 2 else 0
                            latin_products[code] = (
                                latin_products.get(code, 0) + stock
                            )
                            latin_raw_list.append(
                                {
                                    "row_id": str(uuid.uuid4()),
                                    "product_code": code,
                                    "description": str(row[1]).strip()
                                    if len(row) > 1 and row[1] is not None
                                    else "",
                                    "total_stock": stock,
                                    "available_qty": _safe_int(row[4])
                                    if len(row) > 4
                                    else 0,
                                }
                            )
            SharedStore.set_latin_raw_data(latin_raw_list)
            self.latin_total_products = len(latin_products)
            new_items = []
            now_str = datetime.now().isoformat()
            if len(wb.sheetnames) >= 2:
                vs_sheet = wb.worksheets[1]
                vs_items = []
                for row in vs_sheet.iter_rows(min_row=4, values_only=True):
                    if row and len(row) > 8 and (row[8] is not None):
                        ref = str(row[8]).strip()
                        if not ref:
                            continue
                        latin_stock = _safe_int(row[16]) if len(row) > 16 else 0
                        imc_stock = _safe_int(row[17]) if len(row) > 17 else 0
                        diferencia = _safe_int(row[19]) if len(row) > 19 else 0
                        obs = (
                            str(row[15] or "").strip() if len(row) > 15 else ""
                        )
                        estado = (
                            "Coincide"
                            if diferencia == 0
                            else "Sobrante"
                            if diferencia > 0
                            else "Faltante"
                        )
                        vs_items.append(
                            {
                                "item_id": str(uuid.uuid4()),
                                "product_code": ref,
                                "stock_latin": latin_stock,
                                "stock_imc": imc_stock,
                                "diferencia": diferencia,
                                "estado": estado,
                                "observacion": obs,
                                "imported_at": now_str,
                            }
                        )
                vs_refs = set((item["product_code"] for item in vs_items))
                imc_only = set(imc_pivot.keys()) - vs_refs
                for ref in sorted(imc_only):
                    stock = imc_pivot[ref]
                    vs_items.append(
                        {
                            "item_id": str(uuid.uuid4()),
                            "product_code": ref,
                            "stock_latin": 0,
                            "stock_imc": stock,
                            "diferencia": -stock,
                            "estado": "Faltante",
                            "observacion": "Solo en IMC",
                            "imported_at": now_str,
                        }
                    )
                latin_only = set(latin_products.keys()) - vs_refs
                for ref in sorted(latin_only):
                    stock = latin_products[ref]
                    vs_items.append(
                        {
                            "item_id": str(uuid.uuid4()),
                            "product_code": ref,
                            "stock_latin": stock,
                            "stock_imc": 0,
                            "diferencia": stock,
                            "estado": "Sobrante",
                            "observacion": "Solo en LATIN",
                            "imported_at": now_str,
                        }
                    )
                new_items = vs_items
            else:
                all_refs = set(imc_pivot.keys()) | set(latin_products.keys())
                for ref in sorted(all_refs):
                    l_stock = latin_products.get(ref, 0)
                    i_stock = imc_pivot.get(ref, 0)
                    diff = l_stock - i_stock
                    est = (
                        "Coincide"
                        if diff == 0
                        else "Sobrante"
                        if diff > 0
                        else "Faltante"
                    )
                    new_items.append(
                        {
                            "item_id": str(uuid.uuid4()),
                            "product_code": ref,
                            "stock_latin": l_stock,
                            "stock_imc": i_stock,
                            "diferencia": diff,
                            "estado": est,
                            "observacion": "Generado por unión",
                            "imported_at": now_str,
                        }
                    )
            wb.close()
            if new_items:
                SharedStore.set_inventory_items(new_items)
                self.inventory_items = new_items
                self._compute_stats()
                try:
                    from openpyxl.styles import Font, PatternFill

                    wb_report = openpyxl.load_workbook(original_path)
                    if "DIFERENCIAS" in wb_report.sheetnames:
                        del wb_report["DIFERENCIAS"]
                    ws = wb_report.create_sheet("DIFERENCIAS")
                    ws.append(
                        [
                            "CÓDIGO",
                            "STOCK LATIN",
                            "STOCK IMC",
                            "DIFERENCIA",
                            "ESTADO",
                            "OBSERVACIÓN",
                        ]
                    )
                    header_font = Font(bold=True, color="FFFFFF")
                    header_fill = PatternFill(
                        start_color="DC2626",
                        end_color="DC2626",
                        fill_type="solid",
                    )
                    for cell in ws[1]:
                        cell.font = header_font
                        cell.fill = header_fill
                    for item in self.inventory_items:
                        if item.get("diferencia", 0) != 0:
                            ws.append(
                                [
                                    item.get("product_code", ""),
                                    item.get("stock_latin", 0),
                                    item.get("stock_imc", 0),
                                    item.get("diferencia", 0),
                                    item.get("estado", ""),
                                    item.get("observacion", ""),
                                ]
                            )
                    output_path = upload_dir / "reporte_inventario.xlsx"
                    wb_report.save(output_path)
                    wb_report.close()
                except Exception as e:
                    logging.exception(f"Error generating report: {e}")
                self.import_message = (
                    f"{len(new_items)} productos procesados correctamente."
                )
                yield rx.toast(self.import_message)
                yield InventoryState.run_audit()
            else:
                self.import_message = (
                    "No se encontraron productos para procesar."
                )
                yield rx.toast(self.import_message)
        except Exception as e:
            logging.exception(f"Error importando inventario: {e}")
            self.import_message = "Error al procesar el archivo Excel."
            yield rx.toast(self.import_message)
        finally:
            self.is_importing = False

    @rx.event
    async def run_audit(self):
        auth = await self.get_state(AuthState)
        audit = {
            "audit_id": str(uuid.uuid4()),
            "date": datetime.now().strftime("%Y-%m-%d %H:%M"),
            "status": "Perfecta"
            if self.accuracy_pct == 100
            else "Con Discrepancias",
            "total_items": self.total_items,
            "matched": self.matched_count,
            "discrepancies": self.discrepancy_count,
            "accuracy": self.accuracy_pct,
            "created_by": auth.current_user if auth.current_user else "Sistema",
            "notes": "",
        }
        SharedStore.add_inventory_audit(audit)
        self.inventory_audits = SharedStore.get_inventory_audits()
        yield rx.toast("Auditoría generada con éxito.")

    @rx.event
    async def delete_audit(self, audit_id: str):
        auth = await self.get_state(AuthState)
        if auth.current_role != "admin":
            yield rx.toast("No tienes permisos para eliminar auditorías.")
            return
        SharedStore.remove_inventory_audit(audit_id)
        self.inventory_audits = SharedStore.get_inventory_audits()
        yield rx.toast("Auditoría eliminada.")

    @rx.event
    async def delete_inventory_item(self, item_id: str):
        auth = await self.get_state(AuthState)
        if auth.current_role != "admin":
            yield rx.toast("No tienes permisos para eliminar ítems.")
            return
        SharedStore.remove_inventory_item(item_id)
        self.inventory_items = [
            i for i in self.inventory_items if i.get("item_id") != item_id
        ]
        self._compute_stats()
        yield rx.toast("Ítem eliminado.")

    @rx.event
    async def download_report(self):
        upload_dir = rx.get_upload_dir()
        output_path = upload_dir / "reporte_inventario.xlsx"
        if not output_path.exists():
            yield rx.toast("No hay reporte. Reimporta el Excel primero.")
            return
        yield rx.download(
            url="/_upload/reporte_inventario.xlsx",
            filename="Reporte_Cruce_Inventario.xlsx",
        )